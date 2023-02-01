import os
from sqlalchemy import create_engine
import pandas as pd
import openpyxl
import xlrd
import warnings

warnings.filterwarnings("ignore")


db_connection = create_engine('mysql+mysqldb://root:Lokesh@localhost/')


def create_tables():
    db_connection.execute('DROP TABLE IF EXISTS generation_mastersheet')

    db_connection.execute('DROP TABLE IF EXISTS breakdown_mastersheet')

    db_connection.execute('CREATE TABLE IF NOT EXISTS generation_mastersheet(' +
                          'date DATE,' +
                          'financial_year VARCHAR(10),' +
                          'customer_name VARCHAR(255),' +
                          'client_name VARCHAR(255),' +
                          'state VARCHAR(255),' +
                          'site_name VARCHAR(255),' +
                          'wind_turbine_location_number VARCHAR(20),' +
                          'day_generation_kwh FLOAT,' +
                          'day_generation_hours FLOAT,' +
                          'operating_hours FLOAT,' +
                          'machine_availability_percent FLOAT,' +
                          'internal_grid_availability_percent FLOAT,' +
                          'external_grid_availability_percent FLOAT,' +
                          'plant_load_factor FLOAT,' +
                          'nor FLOAT,' +
                          'force_majeure FLOAT,' +
                          'grid_failure FLOAT,' +
                          'internal_grid_failure_hours FLOAT,' +
                          'unscheduled_services FLOAT,' +
                          'scheduled_services FLOAT)')

    db_connection.execute('CREATE TABLE IF NOT EXISTS breakdown_mastersheet(' +
                          'date DATE,' +
                          'financial_year VARCHAR(10),' +
                          'customer_name VARCHAR(255),' +
                          'client_name VARCHAR(255),' +
                          'state VARCHAR(255),' +
                          'site_name VARCHAR(255),' +
                          'wind_turbine_location_number VARCHAR(20),' +
                          'breakdown_remark VARCHAR(500),' +
                          'breakdown_hours FLOAT)')


def timestamp_to_hours(sheet, from_column, to_column):
    for x in range(len(sheet)):
        if len(str(sheet.loc[x, from_column])) > 8:
            sheet.loc[x, to_column] = 24.0
        elif len(str(sheet.loc[x, from_column])) == 8:
            sheet.loc[x, to_column] = float(str(sheet.loc[x, from_column])[-8: -3].replace(':', '.'))
        else:
            sheet.loc[x, to_column] = 0.0

    return sheet


def timestamp_to_hours2(sheet, column):
    for x in range(len(sheet)):
        sheet.loc[x, column] = float(str(sheet.loc[x, column]).replace(':', '.'))
    return sheet


def get_inox_data(gen_data, bd_data):
    input_path = 'DGR Formats/INOX/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2] if filename.endswith('.xlsx')]

    for file in files:
        workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
        if file[1].count('Generation') == 1:
            generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0])

            generation = generation.drop(index=len(generation)-1)

            generation['Date'] = pd.to_datetime(generation['Date'])
            generation['date'] = generation['Date'].dt.date
            generation['financial_year'] = [
                'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(
                    y - 2000)
                for x, y in
                zip(generation['Date'].dt.month, generation['Date'].dt.year)]
            generation['client_name'] = ['INOX'] * len(generation)

            generation = generation.reset_index().rename(columns={'Customer': 'customer_name',
                                                                  'States': 'state', 'Site': 'site_name',
                                                                  'Location No.': 'wind_turbine_location_number',
                                                                  'KWH': 'day_generation_kwh',
                                                                  'MA %': 'machine_availability_percent',
                                                                  'PLF %': 'plant_load_factor',
                                                                  'Production Hrs': 'day_generation_hours',
                                                                  'Operating Hrs': 'operating_hours'})

            generation = timestamp_to_hours2(generation, 'day_generation_hours')
            generation = timestamp_to_hours2(generation, 'operating_hours')

            generation['nor'] = [0.0] * len(generation)
            generation['force_majeure'] = [0.0] * len(generation)
            generation['internal_grid_failure_hours'] = [0.0] * len(generation)
            generation['grid_failure'] = [0.0] * len(generation)
            generation['unscheduled_services'] = [0.0] * len(generation)

            generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                     'state', 'site_name', 'wind_turbine_location_number',
                                     'day_generation_kwh', 'day_generation_hours',
                                     'operating_hours', 'machine_availability_percent',
                                     'plant_load_factor', 'nor', 'force_majeure', 'internal_grid_failure_hours',
                                     'grid_failure', 'unscheduled_services']]

            gen_data = pd.concat([gen_data, generation], ignore_index=True)

        else:
            breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0])

            breakdown['Date'] = pd.to_datetime(breakdown['Date'])
            breakdown['date'] = breakdown['Date'].dt.date
            breakdown['financial_year'] = [
                'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(
                    y - 2000)
                for x, y in
                zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
            breakdown['client_name'] = ['INOX'] * len(breakdown)

            breakdown = timestamp_to_hours(breakdown, 'Total Stop', 'breakdown_hours')

            breakdown = breakdown.reset_index().rename(columns={'Customer': 'customer_name',
                                                                'State': 'state', 'Site': 'site_name',
                                                                'Location No.': 'wind_turbine_location_number',
                                                                'Reason': 'breakdown_remark'})

            bd_factors = {'BOC': 'nor', 'FORCE MAJEURE': 'force_majeure', 'GA INT': 'internal_grid_failure_hours',
                          'GA EXT': 'grid_failure', 'WTG': 'unscheduled_services'}

            for x in range(len(breakdown)):
                if not breakdown.loc[x, 'Stop Due To'] == 'ACTIVITY BEYOND O&M CONTRACT':
                    temp = gen_data[gen_data['date'] == breakdown.loc[x, 'date']]
                    temp = temp[
                        gen_data['wind_turbine_location_number'] == breakdown.loc[x, 'wind_turbine_location_number']]
                    gen_data.loc[temp.index.values, bd_factors[breakdown.loc[x, 'Stop Due To']]] += breakdown.loc[
                        x, 'breakdown_hours']

            breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                                   'state', 'site_name', 'wind_turbine_location_number',
                                   'breakdown_remark', 'breakdown_hours']]

            bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_regen_data(gen_data, bd_data):
    input_path = 'DGR Formats/REGEN/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2] if filename.endswith('.xlsx')]

    for file in files:
        workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)

        generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[1])
        breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[4])

        generation = generation[generation['Date'].isna() == False]
        breakdown = breakdown[breakdown['Date'].isna() == False]

        generation['Date'] = pd.to_datetime(generation['Date'])
        generation['date'] = generation['Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(generation['Date'].dt.month, generation['Date'].dt.year)]
        generation['client_name'] = ['ReGen'] * len(generation)

        generation = generation.reset_index().rename(columns={'Customer Name': 'customer_name',
                                                              'State': 'state', 'Site': 'site_name',
                                                              'WEC Name Project Label': 'wind_turbine_location_number',
                                                              'Daily Generation(KWH)': 'day_generation_kwh',
                                                              'Daily Generation(HRS)': 'day_generation_hours',
                                                              'Machine Availability %': 'machine_availability_percent',
                                                              'Int.Grid Avail %': 'internal_grid_availability_percent',
                                                              'Ext.Grid Avail %': 'external_grid_availability_percent',
                                                              'PLF%': 'plant_load_factor',
                                                              'GF': 'grid_failure',
                                                              'Force majeure': 'force_majeure',
                                                              'S': 'scheduled_services',
                                                              'U': 'unscheduled_services',
                                                              'IGF': 'internal_grid_failure_hours'})

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'state', 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'machine_availability_percent', 'internal_grid_availability_percent',
                                 'external_grid_availability_percent', 'plant_load_factor', 'grid_failure',
                                 'force_majeure', 'scheduled_services', 'unscheduled_services',
                                 'internal_grid_failure_hours']]

        breakdown['Date'] = pd.to_datetime(breakdown['Date'])
        breakdown['date'] = breakdown['Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
        breakdown['client_name'] = ['ReGen'] * len(breakdown)
        breakdown['customer_name'] = [generation.loc[0, 'customer_name']] * len(breakdown)
        breakdown['state'] = [generation.loc[0, 'state']] * len(breakdown)
        breakdown['site_name'] = [generation.loc[0, 'site_name']] * len(breakdown)

        breakdown = breakdown.reset_index().rename(columns={'Location': 'wind_turbine_location_number',
                                                            'Details': 'breakdown_remark',
                                                            'Duration': 'breakdown_hours'})

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'state', 'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        all_turbines = list(set(generation['wind_turbine_location_number']))
        temp = pd.DataFrame(columns=breakdown.columns)
        for x in range(len(breakdown)):
            if breakdown.loc[x, 'wind_turbine_location_number'].__contains__(','):
                turbines = list(breakdown.loc[x, 'wind_turbine_location_number'].split(','))
                for turbine in turbines:
                    temp.loc[len(temp)] = breakdown.loc[x]
                    temp.loc[len(temp) - 1, 'wind_turbine_location_number'] = turbine
            elif breakdown.loc[x, 'wind_turbine_location_number'] in ['ALL Locations', 'All Location', 'All WEC', 'Feeder-1 & Feeder-02']:
                for turbine in all_turbines:
                    temp.loc[len(temp)] = breakdown.loc[x]
                    temp.loc[len(temp) - 1, 'wind_turbine_location_number'] = turbine
            elif breakdown.loc[x, 'wind_turbine_location_number'] in ['FEEDER-01', 'Feeder - 01', 'Feeder-01 all locations', 'Feeder-1', 'FD#1']:
                for turbine in ('GREDW-07', 'GREDW-08', 'GREDW-10', 'GREDW-12', 'GREDW-13', 'GREDW-14', 'GREDW-15'):
                    temp.loc[len(temp)] = breakdown.loc[x]
                    temp.loc[len(temp) - 1, 'wind_turbine_location_number'] = turbine
            elif breakdown.loc[x, 'wind_turbine_location_number'] in ['FEEDER-02', 'Feeder - 02', 'Feeder-02 all locations', 'Feeder-2', 'FD#2']:
                for turbine in ('GREDW-01', 'GREDW-02', 'GREDW-03', 'GREDW-04', 'GREDW-05', 'GREDW-06', 'GREDW-09', 'GREDW-11'):
                    temp.loc[len(temp)] = breakdown.loc[x]
                    temp.loc[len(temp) - 1, 'wind_turbine_location_number'] = turbine
        breakdown = pd.concat([breakdown, temp], ignore_index=True)
        breakdown = breakdown[breakdown['wind_turbine_location_number'].isin(all_turbines) == True]

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_senvion_data(gen_data, bd_data):
    input_path = 'DGR Formats/SENVION/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2] if filename.endswith('.xls')]

    for file in files:
        workbook = xlrd.open_workbook(file[0] + '/' + file[1])

        generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0], skiprows=[0])
        breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[1])

        generation['Date'] = pd.to_datetime(generation['Date'])
        generation['date'] = generation['Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(generation['Date'].dt.month, generation['Date'].dt.year)]
        generation['client_name'] = ['SENVION'] * len(generation)
        file_name = file[1].split('-')
        generation['customer_name'] = [file_name[2].split('_')[0]] * len(generation)
        generation['site_name'] = [file_name[1]] * len(generation)

        generation = timestamp_to_hours(generation, 'Production hours', 'day_generation_hours')
        generation = timestamp_to_hours(generation, 'External Grid Down Time', 'grid_failure')
        generation = timestamp_to_hours(generation, 'Stoppage due to Customer / Utility Account/ Force major', 'force_majeure')
        generation = timestamp_to_hours(generation, 'Scheduled services', 'scheduled_services')
        generation = timestamp_to_hours(generation, 'Unscheduled services', 'unscheduled_services')
        generation = timestamp_to_hours(generation, 'Enviromental', 'nor')
        generation = timestamp_to_hours(generation, 'Internal Grid Down Time', 'internal_grid_failure_hours')

        generation = generation.reset_index().rename(columns={'LOCATION  NO.': 'wind_turbine_location_number',
                                                              'WTG Production kWh ': 'day_generation_kwh',
                                                              'WTG Availability%': 'machine_availability_percent',
                                                              'Int. Grid Availability%': 'internal_grid_availability_percent',
                                                              'Ext. Grid Availability%': 'external_grid_availability_percent',
                                                              'PLF (WTG efficiency) %': 'plant_load_factor'})

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'machine_availability_percent', 'internal_grid_availability_percent',
                                 'external_grid_availability_percent', 'plant_load_factor',
                                 'grid_failure', 'force_majeure', 'nor', 'internal_grid_failure_hours',
                                 'scheduled_services', 'unscheduled_services']]

        breakdown['Date'] = pd.to_datetime(breakdown['Date'])
        breakdown['date'] = breakdown['Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
        breakdown['client_name'] = ['Senvion'] * len(breakdown)
        breakdown['customer_name'] = [file_name[2].split('_')[0]] * len(breakdown)
        breakdown['site_name'] = [file_name[1]] * len(breakdown)

        breakdown = timestamp_to_hours(breakdown, 'Total Duration', 'breakdown_hours')

        breakdown = breakdown.reset_index().rename(columns={'Loc No': 'wind_turbine_location_number',
                                                            'Error Description': 'breakdown_remark'})

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        all_turbines = list(set(generation['wind_turbine_location_number']))
        temp = pd.DataFrame(columns=breakdown.columns)
        for x in range(len(breakdown)):
            if breakdown.loc[x, 'wind_turbine_location_number'] == 'ALL WTG':
                for turbine in all_turbines:
                    temp.loc[len(temp)] = breakdown.loc[x]
                    temp.loc[len(temp) - 1, 'wind_turbine_location_number'] = turbine
        breakdown = pd.concat([breakdown, temp], ignore_index=True)
        breakdown = breakdown[breakdown['wind_turbine_location_number'] != 'ALL WTG']

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_suzlon_data(gen_data, bd_data):
    input_path = 'DGR Formats/SUZLON/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2] if filename.endswith('.xlsx')]

    for file in files:
        workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)

        generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0])
        breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[1])

        generation['Gen. Date'] = pd.to_datetime(generation['Gen. Date'])
        generation['date'] = generation['Gen. Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(generation['Gen. Date'].dt.month, generation['Gen. Date'].dt.year)]
        generation['client_name'] = ['Suzlon'] * len(generation)

        generation.replace('*', 0.0, inplace=True)
        generation.replace('**', 0.0, inplace=True)

        generation = generation.reset_index().rename(columns={'Customer Name': 'customer_name',
                                                              'State': 'state', 'Site': 'site_name',
                                                              'Loc. No.': 'wind_turbine_location_number',
                                                              'Gen. (kwh) DAY': 'day_generation_kwh',
                                                              'Gen Hrs.': 'day_generation_hours',
                                                              'Opr Hrs.': 'operating_hours',
                                                              'M/C Avail.%': 'machine_availability_percent',
                                                              '%PLF DAY': 'plant_load_factor',
                                                              'GF': 'grid_failure',
                                                              'FM': 'force_majeure',
                                                              'S': 'scheduled_services',
                                                              'U': 'unscheduled_services'})

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'state', 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'operating_hours', 'machine_availability_percent',
                                 'plant_load_factor', 'grid_failure', 'force_majeure',
                                 'scheduled_services', 'unscheduled_services']]

        breakdown['Gen. Date'] = pd.to_datetime(breakdown['Gen. Date'])
        breakdown['date'] = breakdown['Gen. Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(breakdown['Gen. Date'].dt.month, breakdown['Gen. Date'].dt.year)]
        breakdown['client_name'] = ['Suzlon'] * len(breakdown)

        breakdown = breakdown.reset_index().rename(columns={'Customer Name': 'customer_name',
                                                            'State': 'state', 'Site': 'site_name',
                                                            'Loc. No.': 'wind_turbine_location_number',
                                                            'Breakdown Remark': 'breakdown_remark',
                                                            'Breakdown Hrs.': 'breakdown_hours'})

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'state', 'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_tswind_data(gen_data, bd_data):
    input_path = 'DGR Formats/TS WIND/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2] if filename.endswith('.xlsx')]

    for file in files:
        workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
        sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0, 1, 3])
        sheet['DATE'] = pd.to_datetime(sheet['DATE'])
        sheet['date'] = sheet['DATE'].dt.date
        sheet['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(sheet['DATE'].dt.month, sheet['DATE'].dt.year)]
        sheet['client_name'] = ['TS Wind'] * len(sheet)

        sheet = timestamp_to_hours(sheet, 'PROD.', 'day_generation_hours')
        sheet = timestamp_to_hours(sheet, 'GRID DROP', 'grid_failure')
        sheet = timestamp_to_hours(sheet, 'FM.', 'force_majeure')
        sheet = timestamp_to_hours(sheet, 'MAINT.', 'scheduled_services')
        sheet = timestamp_to_hours(sheet, 'ERROR', 'unscheduled_services')

        sheet = sheet.reset_index().rename(columns={'CUSTOMER': 'customer_name', 'SITE': 'site_name',
                                                    'WTG': 'wind_turbine_location_number',
                                                    'GEN(FTD)': 'day_generation_kwh',
                                                    'MA(FTD)': 'machine_availability_percent',
                                                    'PLF': 'plant_load_factor', 'ERROR DETAILS': 'breakdown_remark'})

        generation = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                            'site_name', 'wind_turbine_location_number',
                            'day_generation_kwh', 'day_generation_hours',
                            'machine_availability_percent', 'plant_load_factor',
                            'grid_failure', 'force_majeure', 'scheduled_services', 'unscheduled_services']]

        breakdown = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                           'site_name', 'wind_turbine_location_number', 'breakdown_remark']]

        breakdown.dropna(inplace=True)

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_windworld_data(gen_data, bd_data):
    input_path = 'DGR Formats/WIND WORLD/'

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2]
             if filename.endswith('.xlsx') or filename.endswith('.xls')]

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0], skiprows=[0])

        sheet['DATE'] = pd.to_datetime(sheet['DATE'])
        sheet['date'] = sheet['DATE'].dt.date
        sheet['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(sheet['DATE'].dt.month, sheet['DATE'].dt.year)]
        sheet['client_name'] = ['Wind World'] * len(sheet)

        sheet = timestamp_to_hours(sheet, 'O.Hrs', 'operating_hours')

        bd_factors = {'BM': 'unscheduled_services', 'BD': 'unscheduled_services', 'GF': 'grid_failure',
                      'GS': 'grid_failure', 'PM': 'schedules_services', 'SD': 'schedules_services',
                      'FM': 'force_majeure'}

        sheet['grid_failure'] = [0] * len(sheet)
        sheet['force_majeure'] = [0] * len(sheet)
        sheet['scheduled_services'] = [0] * len(sheet)
        sheet['unscheduled_services'] = [0] * len(sheet)

        sheet.fillna('', inplace=True)
        for x in range(len(sheet)):
            if sheet.loc[x, 'REMARKS'] != '' and bd_data.get(sheet.loc[x, 'REMARKS'][:2]) is not None:
                sheet.loc[x, bd_factors[sheet.loc[x, 'REMARKS'][:2]]] = 1

        sheet = sheet.reset_index().rename(columns={'Customer': 'customer_name',
                                                    'STATE ': 'state', 'SITE': 'site_name',
                                                    'WEC': 'wind_turbine_location_number',
                                                    'GENERATION': 'day_generation_kwh',
                                                    'MA ': 'machine_availability_percent',
                                                    'GIA': 'internal_grid_availability_percent',
                                                    'CF': 'plant_load_factor', 'REMARKS': 'breakdown_remark'})

        generation = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                            'state', 'site_name', 'wind_turbine_location_number',
                            'day_generation_kwh', 'operating_hours',
                            'machine_availability_percent', 'internal_grid_availability_percent',
                            'plant_load_factor', 'grid_failure', 'force_majeure',
                            'scheduled_services', 'unscheduled_services']]

        breakdown = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                           'state', 'site_name', 'wind_turbine_location_number',
                           'breakdown_remark']]

        breakdown = breakdown[breakdown['breakdown_remark'] != '']

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def consolidate():

    create_tables()

    gen_data = pd.DataFrame()
    bd_data = pd.DataFrame()

    gen_data, bd_data = get_inox_data(gen_data, bd_data)

    gen_data, bd_data = get_regen_data(gen_data, bd_data)

    gen_data, bd_data = get_senvion_data(gen_data, bd_data)

    gen_data, bd_data = get_suzlon_data(gen_data, bd_data)

    gen_data, bd_data = get_tswind_data(gen_data, bd_data)

    gen_data, bd_data = get_windworld_data(gen_data, bd_data)

    customers = {'D.J. Malpani': 'DJM', 'D J MALPANI': 'DJM', 'Giriraj Enterprises': 'GE', 'DJM': 'DJM',
                 'D J Malpani': 'DJM', 'NAKODA MACHINERY PVT. LTD.': 'NMPL', 'DJ Malpani': 'DJM', 'DJ Malpani - Palakkad': 'DJM',
                 'DJ Malpani - Sadla': 'DJM', 'DJ Malpani - Savarkundla': 'DJM', 'Giriraj Enterprises - Bagewadi': 'GE',
                 'IVY Ecoenergy India Private Ltd': 'IVY Ecoenergy India Private Ltd'}
    gen_data['customer_name'] = [customers[x] for x in gen_data['customer_name']]
    bd_data['customer_name'] = [customers[x] for x in bd_data['customer_name']]

    gen_data.to_sql(con=db_connection, name='generation_mastersheet', if_exists='append', index=False)
    bd_data.to_sql(con=db_connection, name='breakdown_mastersheet', if_exists='append', index=False)

    print('DGR Consolidated Successfully')
    print('-----------------------------', end='\n\n')
